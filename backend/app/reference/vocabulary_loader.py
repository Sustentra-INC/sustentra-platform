from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
from typing import Any


class VocabularyLoaderError(RuntimeError):
    """Base error for vocabulary loader failures."""


class VocabularyWorkbookNotFoundError(VocabularyLoaderError):
    """Raised when the vocabulary workbook file cannot be located."""


class VocabularySheetError(VocabularyLoaderError):
    """Raised when required workbook sheets are missing."""


class VocabularyColumnError(VocabularyLoaderError):
    """Raised when required columns are missing from a sheet."""


class VocabularyDataError(VocabularyLoaderError):
    """Raised when workbook row values fail required parsing/validation."""


@dataclass(frozen=True)
class VocabularyCanonicalType:
    canonical_type_id: str
    canonical_name: str
    definition: str
    data_category: str
    ghg_scope: str | None
    content_inventory_coarse: str | None
    common_aliases: tuple[str, ...]
    source_basis: str | None
    population_status: str
    source_version: str
    last_reviewed: str | None


@dataclass(frozen=True)
class VocabularyVariant:
    variant_id: str
    canonical_type_id: str
    variant_archetype: str
    issuer_origin: str | None
    jurisdiction_region: str | None
    filename_patterns: tuple[str, ...]
    layout_features: tuple[str, ...]
    header_terms: tuple[str, ...]
    key_phrases: tuple[str, ...]
    multi_type: bool
    confidence_threshold_default: float
    confidence_threshold_raw: str | float | None
    calibration_status: str
    source_exemplar: str | None
    population_status: str


@dataclass(frozen=True)
class VocabularyLibrary:
    canonical_types: tuple[VocabularyCanonicalType, ...]
    variants: tuple[VocabularyVariant, ...]
    canonical_by_id: dict[str, VocabularyCanonicalType]
    variant_by_id: dict[str, VocabularyVariant]
    variants_by_canonical_type_id: dict[str, tuple[VocabularyVariant, ...]]


REQUIRED_SHEETS = ("Canonical_Types", "Variants")

CANONICAL_REQUIRED_COLUMNS = (
    "canonical_type_id",
    "canonical_name",
    "definition",
    "data_category",
    "ghg_scope",
    "content_inventory_coarse",
    "common_aliases",
    "source_basis",
    "population_status",
    "source_version",
    "last_reviewed",
)

VARIANT_REQUIRED_COLUMNS = (
    "variant_id",
    "canonical_type_id",
    "variant_archetype",
    "issuer_origin",
    "jurisdiction_region",
    "filename_patterns",
    "layout_features",
    "header_terms",
    "key_phrases",
    "multi_type",
    "confidence_threshold_default",
    "calibration_status",
    "source_exemplar",
    "population_status",
)

PIPE_DELIMITED_COLUMNS = {
    "common_aliases",
    "filename_patterns",
    "layout_features",
    "header_terms",
    "key_phrases",
}

TRUE_VALUES = {"true", "yes", "1", "y", "t"}
FALSE_VALUES = {"false", "no", "0", "n", "f"}

CONFIDENCE_PREFIX_PATTERN = re.compile(r"^\s*([0-9]+(?:\.[0-9]+)?)")


def find_default_vocabulary_workbook(repo_root: Path | None = None) -> Path:
    """Find the default vocabulary workbook path under reference-data."""

    root = repo_root.resolve() if repo_root is not None else Path(__file__).resolve().parents[3]
    reference_root = root / "reference-data"
    if not reference_root.exists():
        raise VocabularyWorkbookNotFoundError(
            f"reference-data directory was not found at: {reference_root}"
        )

    preferred_dirs = [
        reference_root / "vocab-library",
        reference_root / "vocab library",
        reference_root / "vocab_library",
        reference_root,
    ]

    matches: dict[str, Path] = {}
    for directory in preferred_dirs:
        if not directory.exists() or not directory.is_dir():
            continue
        for workbook in sorted(directory.glob("Vocabulary_Library*.xlsx")):
            if workbook.is_file():
                matches[str(workbook.resolve())] = workbook.resolve()

    if not matches:
        for workbook in sorted(reference_root.rglob("Vocabulary_Library*.xlsx")):
            if workbook.is_file():
                matches[str(workbook.resolve())] = workbook.resolve()

    if not matches:
        raise VocabularyWorkbookNotFoundError(
            "Could not find a workbook matching Vocabulary_Library*.xlsx under reference-data."
        )

    found = list(matches.values())
    exact = [path for path in found if path.name == "Vocabulary_Library_v1.0.xlsx"]
    return sorted(exact or found)[0]


def load_default_vocabulary_library(repo_root: Path | None = None) -> VocabularyLibrary:
    """Load vocabulary library from default reference-data location."""

    workbook_path = find_default_vocabulary_workbook(repo_root=repo_root)
    return load_vocabulary_library(workbook_path)


def load_vocabulary_library(workbook_path: Path) -> VocabularyLibrary:
    """Load and validate vocabulary workbook rows into typed structures."""

    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - environment dependency
        raise VocabularyLoaderError(
            "openpyxl is required to load the vocabulary workbook."
        ) from exc

    if not workbook_path.exists() or not workbook_path.is_file():
        raise VocabularyWorkbookNotFoundError(
            f"Vocabulary workbook not found at: {workbook_path}"
        )

    workbook = load_workbook(workbook_path, data_only=True)
    missing_sheets = [name for name in REQUIRED_SHEETS if name not in workbook.sheetnames]
    if missing_sheets:
        raise VocabularySheetError(
            f"Vocabulary workbook missing required sheet(s): {missing_sheets}. "
            f"Found sheets: {workbook.sheetnames}"
        )

    canonical_rows = _read_sheet_rows(
        workbook["Canonical_Types"],
        required_columns=CANONICAL_REQUIRED_COLUMNS,
        sheet_name="Canonical_Types",
    )
    variant_rows = _read_sheet_rows(
        workbook["Variants"],
        required_columns=VARIANT_REQUIRED_COLUMNS,
        sheet_name="Variants",
    )

    canonical_types: list[VocabularyCanonicalType] = []
    canonical_by_id: dict[str, VocabularyCanonicalType] = {}

    for row_number, row in canonical_rows:
        canonical_type_id = _required_text(row, "canonical_type_id", "Canonical_Types", row_number)
        if canonical_type_id in canonical_by_id:
            raise VocabularyDataError(
                f"Canonical_Types row {row_number}: duplicate canonical_type_id '{canonical_type_id}'."
            )

        canonical_type = VocabularyCanonicalType(
            canonical_type_id=canonical_type_id,
            canonical_name=_required_text(row, "canonical_name", "Canonical_Types", row_number),
            definition=_required_text(row, "definition", "Canonical_Types", row_number),
            data_category=_required_text(row, "data_category", "Canonical_Types", row_number),
            ghg_scope=_optional_text(row.get("ghg_scope")),
            content_inventory_coarse=_optional_text(row.get("content_inventory_coarse")),
            common_aliases=_parse_pipe_delimited(row.get("common_aliases")),
            source_basis=_optional_text(row.get("source_basis")),
            population_status=_required_text(row, "population_status", "Canonical_Types", row_number),
            source_version=_required_text(row, "source_version", "Canonical_Types", row_number),
            last_reviewed=_optional_text(row.get("last_reviewed")),
        )
        canonical_types.append(canonical_type)
        canonical_by_id[canonical_type_id] = canonical_type

    variants: list[VocabularyVariant] = []
    variant_by_id: dict[str, VocabularyVariant] = {}

    for row_number, row in variant_rows:
        variant_id = _required_text(row, "variant_id", "Variants", row_number)
        if variant_id in variant_by_id:
            raise VocabularyDataError(
                f"Variants row {row_number}: duplicate variant_id '{variant_id}'."
            )

        canonical_type_id = _required_text(row, "canonical_type_id", "Variants", row_number)
        if canonical_type_id not in canonical_by_id:
            raise VocabularyDataError(
                "Variants row {row}: canonical_type_id '{canonical}' does not exist in "
                "Canonical_Types.".format(row=row_number, canonical=canonical_type_id)
            )

        threshold, threshold_raw = _parse_confidence_threshold(
            row.get("confidence_threshold_default"),
            "Variants",
            row_number,
            "confidence_threshold_default",
        )

        variant = VocabularyVariant(
            variant_id=variant_id,
            canonical_type_id=canonical_type_id,
            variant_archetype=_required_text(row, "variant_archetype", "Variants", row_number),
            issuer_origin=_optional_text(row.get("issuer_origin")),
            jurisdiction_region=_optional_text(row.get("jurisdiction_region")),
            filename_patterns=_parse_pipe_delimited(row.get("filename_patterns")),
            layout_features=_parse_pipe_delimited(row.get("layout_features")),
            header_terms=_parse_pipe_delimited(row.get("header_terms")),
            key_phrases=_parse_pipe_delimited(row.get("key_phrases")),
            multi_type=_parse_boolean(row.get("multi_type"), "Variants", row_number, "multi_type"),
            confidence_threshold_default=threshold,
            confidence_threshold_raw=threshold_raw,
            calibration_status=_required_text(row, "calibration_status", "Variants", row_number),
            source_exemplar=_optional_text(row.get("source_exemplar")),
            population_status=_required_text(row, "population_status", "Variants", row_number),
        )
        variants.append(variant)
        variant_by_id[variant_id] = variant

    variants_by_canonical: dict[str, list[VocabularyVariant]] = {
        canonical_id: [] for canonical_id in canonical_by_id
    }
    for variant in variants:
        variants_by_canonical.setdefault(variant.canonical_type_id, []).append(variant)

    return VocabularyLibrary(
        canonical_types=tuple(canonical_types),
        variants=tuple(variants),
        canonical_by_id=canonical_by_id,
        variant_by_id=variant_by_id,
        variants_by_canonical_type_id={
            canonical_id: tuple(items)
            for canonical_id, items in variants_by_canonical.items()
        },
    )


def _read_sheet_rows(
    worksheet: Any,
    required_columns: tuple[str, ...],
    sheet_name: str,
) -> list[tuple[int, dict[str, Any]]]:
    rows = worksheet.iter_rows(values_only=True)
    header_row = next(rows, None)
    if header_row is None:
        raise VocabularyColumnError(f"{sheet_name} sheet is empty.")

    headers = [_optional_text(value) or "" for value in header_row]
    column_positions = {
        header: index
        for index, header in enumerate(headers)
        if header
    }

    missing_columns = [name for name in required_columns if name not in column_positions]
    if missing_columns:
        raise VocabularyColumnError(
            f"{sheet_name} sheet missing required column(s): {missing_columns}."
        )

    parsed_rows: list[tuple[int, dict[str, Any]]] = []
    for row_number, values in enumerate(rows, start=2):
        if _row_is_empty(values):
            continue
        row_data = {
            column_name: values[position] if position < len(values) else None
            for column_name, position in column_positions.items()
        }
        parsed_rows.append((row_number, row_data))

    return parsed_rows


def _row_is_empty(values: tuple[Any, ...]) -> bool:
    for value in values:
        if _optional_text(value) is not None:
            return False
    return True


def _required_text(row: dict[str, Any], key: str, sheet_name: str, row_number: int) -> str:
    value = _optional_text(row.get(key))
    if value is None:
        raise VocabularyDataError(
            f"{sheet_name} row {row_number}: '{key}' is required and cannot be empty."
        )
    return value


def _optional_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _parse_pipe_delimited(value: Any) -> tuple[str, ...]:
    text = _optional_text(value)
    if text is None:
        return ()

    if "|" not in text:
        return (text,)

    terms = [item.strip() for item in text.split("|")]
    return tuple(item for item in terms if item)


def _parse_boolean(value: Any, sheet_name: str, row_number: int, key: str) -> bool:
    if isinstance(value, bool):
        return value

    if isinstance(value, (int, float)) and value in (0, 1):
        return bool(int(value))

    text = _optional_text(value)
    if text is None:
        raise VocabularyDataError(
            f"{sheet_name} row {row_number}: '{key}' is required and must be boolean-like."
        )

    normalized = text.lower()
    if normalized in TRUE_VALUES:
        return True
    if normalized in FALSE_VALUES:
        return False

    raise VocabularyDataError(
        f"{sheet_name} row {row_number}: '{key}' value '{text}' is not a valid boolean."
    )


def _parse_confidence_threshold(
    value: Any,
    sheet_name: str,
    row_number: int,
    key: str,
) -> tuple[float, str | float | None]:
    if value is None or (isinstance(value, str) and not value.strip()):
        raise VocabularyDataError(
            f"{sheet_name} row {row_number}: '{key}' is required and cannot be empty."
        )

    threshold_raw: str | float | None
    if isinstance(value, (int, float)):
        threshold = float(value)
        threshold_raw = float(value)
    else:
        text = _optional_text(value)
        if text is None:
            raise VocabularyDataError(
                f"{sheet_name} row {row_number}: '{key}' is required and cannot be empty."
            )
        match = CONFIDENCE_PREFIX_PATTERN.match(text)
        if not match:
            raise VocabularyDataError(
                f"{sheet_name} row {row_number}: '{key}' value '{text}' is invalid."
            )
        threshold = float(match.group(1))
        threshold_raw = text

    if threshold < 0 or threshold > 1:
        raise VocabularyDataError(
            f"{sheet_name} row {row_number}: '{key}' value {threshold} must be in [0, 1]."
        )

    return threshold, threshold_raw
