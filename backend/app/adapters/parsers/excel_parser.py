"""Excel workbook parser adapter.

Normalizes ``.xlsx`` workbooks into a ``parser_output`` object using ``openpyxl``.
Every non-empty cell becomes a traceable text block plus source reference, and
each worksheet with content produces a table. No business-field inference is
performed here.
"""

from __future__ import annotations

from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from backend.app.adapters.parsers.base import (
    PARSER_RUNTIME_VERSION,
    build_empty_parser_output,
    build_failed_parser_output,
    build_parser_output,
    build_source_reference,
    build_text_block,
    build_warning,
    snippet,
)

PARSER_NAME = "openpyxl"


def _coerce_cell_value(value: Any) -> str | int | float | bool | None:
    """Coerce a workbook cell value into a JSON/table-safe scalar."""

    if value is None or isinstance(value, bool):
        return value
    if isinstance(value, (int, float, str)):
        return value
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return str(value)


class ExcelParser:
    """Adapter that normalizes ``.xlsx`` workbooks into ``parser_output`` dicts."""

    parser_name = PARSER_NAME
    parser_version = PARSER_RUNTIME_VERSION

    def parse(self, file_path: str | Path, document_id: str, processing_run_id: str) -> dict:
        path = Path(file_path)

        try:
            from openpyxl import load_workbook
        except ImportError:
            return build_failed_parser_output(
                document_id=document_id,
                processing_run_id=processing_run_id,
                parser_name=self.parser_name,
                warnings=[
                    build_warning(
                        "missing_dependency",
                        "openpyxl is not installed; cannot parse Excel workbook.",
                        severity="error",
                    )
                ],
            )

        try:
            workbook = load_workbook(path, data_only=True, read_only=True)
        except Exception as exc:  # noqa: BLE001 - normalize any load failure
            return build_failed_parser_output(
                document_id=document_id,
                processing_run_id=processing_run_id,
                parser_name=self.parser_name,
                warnings=[
                    build_warning(
                        "workbook_load_failed",
                        f"Could not open workbook: {type(exc).__name__}.",
                        severity="error",
                    )
                ],
            )

        text_blocks: list[dict] = []
        source_references: list[dict] = []
        tables: list[dict] = []
        warnings: list[dict] = []
        cell_counter = 0

        try:
            for sheet_index, worksheet in enumerate(workbook.worksheets, start=1):
                sheet_name = worksheet.title
                table_rows: list[list[Any]] = []
                sheet_has_content = False

                for row in worksheet.iter_rows():
                    coerced_row: list[Any] = []
                    for cell in row:
                        coerced = _coerce_cell_value(cell.value)
                        coerced_row.append(coerced)

                        if coerced is None or (isinstance(coerced, str) and not coerced.strip()):
                            continue

                        sheet_has_content = True
                        cell_counter += 1
                        block_id = f"{self.parser_name}-{sheet_name}-{cell.coordinate}"
                        source_reference_id = f"SRC-{self.parser_name}-c{cell_counter}"
                        cell_text = str(coerced)
                        source_references.append(
                            build_source_reference(
                                source_reference_id,
                                document_id,
                                sheet_name=sheet_name,
                                cell_or_range=cell.coordinate,
                                text_snippet=snippet(cell_text),
                                parser_block_ids=[block_id],
                                source_kind="excel_cell",
                            )
                        )
                        text_blocks.append(
                            build_text_block(
                                block_id,
                                cell_text,
                                source_reference_id=source_reference_id,
                            )
                        )

                    table_rows.append(coerced_row)

                if sheet_has_content:
                    tables.append(
                        {
                            "table_id": f"{self.parser_name}-{sheet_name}-t{sheet_index}",
                            "page_number": None,
                            "sheet_name": sheet_name,
                            "rows": table_rows,
                            "confidence": None,
                            "bounding_box": None,
                            "source_reference_id": None,
                        }
                    )
        finally:
            workbook.close()

        if not text_blocks:
            warnings.append(
                build_warning(
                    "empty_document",
                    "Workbook contained no non-empty cells.",
                    severity="info",
                )
            )
            return build_empty_parser_output(
                document_id=document_id,
                processing_run_id=processing_run_id,
                parser_name=self.parser_name,
                warnings=warnings,
            )

        return build_parser_output(
            document_id=document_id,
            processing_run_id=processing_run_id,
            parser_name=self.parser_name,
            status="parsed",
            text_blocks=text_blocks,
            tables=tables,
            source_references=source_references,
            warnings=warnings,
        )
